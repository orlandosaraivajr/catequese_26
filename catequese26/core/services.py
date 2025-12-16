from django.conf import settings
from datetime import date, datetime
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import Frame, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


def data_hoje():
    meses = { 1: "janeiro",2: "fevereiro",3: "março",
    4: "abril",5: "maio", 6: "junho",7: "julho",
    8: "agosto",9: "setembro",10: "outubro",
    11: "novembro",12: "dezembro" }
    hoje = datetime.today()
    return f"{hoje.day} de {meses[hoje.month]} de {hoje.year}"

def gerar_ficha_catequese(ficha):
    img_path = os.path.join(settings.BASE_DIR, 'static', 'pdf', 'cabecalho.png')

    filename = os.path.join(settings.MEDIA_ROOT, f"ficha_catequese_{ficha.id}.pdf")

    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4

    # Inserindo imagem no topo
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 200, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica", 16)
    c.drawString(130, height - 220, f"INSCRIÇÃO PARA CATEQUESE INFANTIL")
    
    # Nome
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 250, f"Nome:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 250, f"{ficha.nome}")

    # Sexo
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 265, f"Sexo:")
    c.setFont("Helvetica", 11)
    if ficha.sexo == 'M':
        c.drawString(150, height - 265, f"Masculino")
    else:
        c.drawString(150, height - 265, f"Feminino")
    
    # Data de Nascimento
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 280, f"Data Nascimento:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 280, f"{ficha.data_nascimento.strftime("%d/%m/%Y")}")
    
    # Naturalidade
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 295, f"Naturalidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 295, f"{ficha.naturalidade}")
    
    # Nome dos Pais
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 310, f"Pai:")
    c.setFont("Helvetica", 11)
    if ficha.nome_pai:
        c.drawString(150, height - 310, f"{ficha.nome_pai}  -  {ficha.celular_pai}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 325, f"Mãe:")
    c.setFont("Helvetica", 11)
    if ficha.nome_mae:
        c.drawString(150, height - 325, f"{ficha.nome_mae}  -  {ficha.celular_mae}")

    # Endereço
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 340, f"Endereço:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 340, f"{ficha.endereco}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 355, f"Cidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 355, f"{ficha.cidade}  -  {ficha.uf}")
    
    # Batizado
    c.setFont("Helvetica-Bold", 11)
    if ficha.batizado:
        c.drawString(50, height - 375, f"Batizado")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 390, f"Data: {ficha.batizado_data.strftime("%d/%m/%Y") if ficha.batizado_data else ''}")
        c.drawString(50, height - 405, f"Diocese: {ficha.batizado_diocese}   Paróquia: {ficha.batizado_paroquia}")
        c.drawString(50, height - 420, f"Celebrante: {ficha.batizado_celebrante} ")
        
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 375, f"Não Batizado")
        c.setFillColor(colors.black)
    
    
    c.setFont("Helvetica", 11)
    c.drawString(50, height - 440, f"Horário:")
    c.setFillColor(colors.blue)
    c.drawString(150, height - 440, f"{ficha.get_horario_display()} ")
    c.setFillColor(colors.black)
    
# Cuidado e Acolhimento da Criança
    c.setFont("Helvetica-Bold", 16)
    c.drawString(170, height - 480, "Cuidado e Acolhimento da Criança")   
    c.setFont("Helvetica", 11)
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        "Para que possamos receber seu filho(a) com todo carinho, atenção e segurança, "
        "pedimos que compartilhe conosco algumas informações importantes. "
        "Tudo será tratado com sigilo e usado apenas para ajudar no bem-estar "
        "da criança durante as atividades.",
        style
    )

    frame = Frame(50, height - 700, 500, 200)  # x, y, largura, altura
    frame.addFromList([paragrafo], c)
    
# 1 Deficiência ou Necessidade Especial
    paragrafo = Paragraph(
        "1. Seu filho(a) possui alguma deficiência que devemos conhecer para acolhê-lo(a) "
        "da melhor forma?",
        style
    )
    frame = Frame(50, height - 750, 500, 200)
    frame.addFromList([paragrafo], c)
    if ficha.possui_deficiencia:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_deficiencia}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)

# 2 Transtornos
    paragrafo = Paragraph(
        "2. Há algum transtorno ou diagnóstico que ajude nossa pastoral a compreender melhor as necessidades da criança?",
        style
    )
    frame = Frame(50, height - 795, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.possui_transtorno:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_transtorno}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 825, 500, 200)
    frame.addFromList([paragrafo], c)
    
# 3 Medicamentos
    paragrafo = Paragraph(
        "3. Seu filho(a) faz uso de algum medicamento contínuo?",
        style
    )
    frame = Frame(50, height - 840, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.medicamento_uso_continuo:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_medicamento}. Horário: {ficha.medicamento_horario}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 855, 500, 200)
    frame.addFromList([paragrafo], c)   
    
# 4 Acompanhamento Psicológico
    paragrafo = Paragraph(
        "4. A criança faz acompanhamento psicológico, psiquiátrico ou terapêutico?",
        style
    )
    frame = Frame(50, height - 870, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.acompanhamento_psicologico:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_acompanhamento}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 885, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            "Agradecemos de coração pela confiança. Nosso desejo é caminhar juntos para que a criança viva cada momento com alegria, segurança e acolhimento.",
        style
    )
    frame = Frame(50, height - 900, 500, 200)
    frame.addFromList([paragrafo], c)
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> Assinatura do Responsável",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    
    
    c.showPage()  # Página 2
    

    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"TERMO DE CONSENTIMENTO PARA TRATAMENTO DE DADOS ")
    c.drawString(120, height - 180, f"PESSOAIS SENSÍVEIS DE CRIANÇAS E ADOLESCENTES")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome_responsavel}, CPF {ficha.cpf_responsavel} na qualidade de responsável legal pelo(a) menor "
        f"{ficha.nome}, por meio deste instrumento, <b>manifesto meu consentimento livre, informado e inequívoco para o tratamento dos dados pessoais do(a) referido"
        "menor</b>, nos seguintes termos, em conformidade com a Lei no 13.709/2018:"
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"1) Autorizo o tratamento a ser realizado pela PARÓQUIA NOSSA SENHORA APARECIDA, pessoa jurídica "
        f"de direito privado, inscrita no CNPJ no 44.802.999/0011-30, situada na Rua dois, no 349, Bairro Aparecida,"
        f"Rio Claro/SP, CEP 13500-270, e e-mail: pnsarc@hotmail.com, doravante denominada CONTROLADORA."
        ,style
    )
    frame = Frame(50, height - 460, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"2) Os dados pessoais do(a) menor serão utilizados exclusivamente para:<br/>"
        f"    I) Inscrição e organização da Catequese;<br/>"
        f"    II) Emitir certificados de conclusão dos sacramentos;<br/>"
        f"    III) Permitir contato com pais e/ou responsáveis.<br/>"
        ,style
    )
    frame = Frame(50, height - 520, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"3) Autorizo o tratamento apenas dos seguintes dados pessoais: nome completo; data de nascimento;"
        f"naturalidade; documento de identificação (RG, CPF, certidão de nascimento ou equivalente); endereço completo; telefone de contato."
        f"Não haverá utilização dos dados para fins comerciais, divulgação indevida ou compartilhamento com terceiros estranhos às atividades religiosas."
        ,style
    )
    frame = Frame(50, height - 590, 500, 200)
    frame.addFromList([paragrafo], c)
    

    paragrafo = Paragraph(
        f"4) Declaro estar ciente de que o armazenamento ocorrerá:"
        f"I) Em fichas físicas e/ou sistemas informatizados da Paróquia;<br/>"
        f"II) Que o acesso será restrito a pessoas autorizadas (secretaria, coordenação da catequese, catequistas e o Pároco, quando necessário);<br/>"
        f"III) Que a CONTROLADORA adotará medidas técnicas e administrativas adequadas para proteger os dados contra acessos não autorizados, perda, divulgação indevida ou qualquer forma de tratamento inadequado.<br/>"
        ,style
    )
    frame = Frame(50, height - 670, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"5) Os dados serão mantidos enquanto o(a) menor estiver matriculado(a) na Catequese, pelo prazo máximo de 3 (três) anos, quando concluída as etapas sacramentais, respeitando-se a necessidade de registros paroquiais.<br/>"
        f"6) Estou ciente de que posso, a qualquer momento solicitar acesso aos dados pessoais do(a) menor; solicitar correção de dados incompletos, inexatos ou desatualizados; requerer a eliminação de dados desnecessários, excessivos ou tratados em desconformidade com a lei.<br/>"
        f"7) Declaro, ainda, estar ciente de que posso revogar este consentimento a qualquer momento, mediante solicitação formal e por escrito à Secretaria Paroquial.<br/>"
        ,style
    )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)
   
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome_responsavel} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)

    c.showPage()  # Página 3 
    
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"AUTORIZAÇÃO PARA USO DE IMAGEM MENORES DE 18 ANOS")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome_responsavel}, CPF {ficha.cpf_responsavel} na qualidade de responsável legal pelo(a) menor "
        f"{ficha.nome}, autorizo, de forma livre, expressa e informada, a Paróquia Nossa Senhora Aparecida, inscrita "
        f"no CNPJ sob o nº 44.802.999/0011-30, a utilizar a imagem, nome e voz do(a) menor, captados em fotografias e/ou vídeos durante atividades e eventos da paróquia, para fins de divulgação em meios impressos, digitais e redes sociais da paróquia, sem qualquer ônus."
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Declaro estar ciente de que a utilização da imagem do(a) menor será feita de acordo com a Lei Geral de Proteção de Dados (Lei nº 13.709/2018 - LGPD), e que posso, a qualquer momento, revogar esta autorização mediante solicitação por escrito à paróquia. "
        ,style
    )
    frame = Frame(50, height - 480, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Estou ciente de que não tenho direito a qualquer remuneração pelo uso da imagem, nome e voz do(a) menor nos termos acima mencionados, e que a presente autorização é concedida por prazo indeterminado, podendo ser revogada a qualquer momento, mediante comunicação por escrito. "
        ,style
    )
    frame = Frame(50, height - 530, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"Por fim, declaro que a presente autorização foi feita de forma livre, sem qualquer coação, e que fui devidamente informado(a) sobre o tratamento dos dados pessoais do(a) menor pela paróquia. "
        ,style
    )
    frame = Frame(50, height - 580, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome_responsavel} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 660, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.save()

    return filename

def gerar_ficha_crisma(ficha):
    anos = ((date.today() - ficha.data_nascimento)/365.2425).days
    if anos < 18:
        return gerar_ficha_crisma_menor(ficha)
    if anos >= 18:
        return gerar_ficha_crisma_maior(ficha)

def gerar_ficha_crisma_menor(ficha):
    img_path = os.path.join(settings.BASE_DIR, 'static', 'pdf', 'cabecalho.png')

    filename = os.path.join(settings.MEDIA_ROOT, f"ficha_crisma_{ficha.id}.pdf")

    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4

    # Inserindo imagem no topo
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 200, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica", 16)
    c.drawString(130, height - 220, f"INSCRIÇÃO PARA CRISMA - ADOLESCENTES")
    
    # Nome
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 250, f"Nome:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 250, f"{ficha.nome}")

    # Sexo
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 265, f"Sexo:")
    c.setFont("Helvetica", 11)
    if ficha.sexo == 'M':
        c.drawString(150, height - 265, f"Masculino")
    else:
        c.drawString(150, height - 265, f"Feminino")
    
    # Data de Nascimento
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 280, f"Data Nascimento:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 280, f"{ficha.data_nascimento.strftime("%d/%m/%Y")}")
    
    # Naturalidade
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 295, f"Naturalidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 295, f"{ficha.naturalidade}")
    
    # Nome dos Pais
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 310, f"Pai:")
    c.setFont("Helvetica", 11)
    if ficha.nome_pai:
        c.drawString(150, height - 310, f"{ficha.nome_pai}  -  {ficha.celular_pai}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 325, f"Mãe:")
    c.setFont("Helvetica", 11)
    if ficha.nome_mae:
        c.drawString(150, height - 325, f"{ficha.nome_mae}  -  {ficha.celular_mae}")

    # Endereço
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 340, f"Endereço:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 340, f"{ficha.endereco}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 355, f"Cidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 355, f"{ficha.cidade}  -  {ficha.uf}")
    
    # Batizado
    c.setFont("Helvetica-Bold", 11)
    if ficha.batizado:
        c.drawString(50, height - 375, f"Batizado")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 390, f"Data: {ficha.batizado_data.strftime("%d/%m/%Y") if ficha.batizado_data else ''}")
        c.drawString(50, height - 405, f"Diocese: {ficha.batizado_diocese}   Paróquia: {ficha.batizado_paroquia}")
        c.drawString(50, height - 420, f"Celebrante: {ficha.batizado_celebrante} ")
        
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 375, f"Não Batizado")
        c.setFillColor(colors.black)
    
    
    c.setFont("Helvetica", 11)
    c.drawString(50, height - 440, f"Horário:")
    c.setFillColor(colors.blue)
    c.drawString(150, height - 440, f"{ficha.get_horario_display()} ")
    c.setFillColor(colors.black)
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

# Cuidado e Acolhimento da Criança
    c.setFont("Helvetica-Bold", 16)
    c.drawString(170, height - 480, "Cuidado e Acolhimento")   
    c.setFont("Helvetica", 11)
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        "Para que possamos receber seu filho(a) com todo carinho, atenção e segurança, "
        "pedimos que compartilhe conosco algumas informações importantes. "
        "Tudo será tratado com sigilo e usado apenas para ajudar no bem-estar "
        "do(a) jovem durante as atividades.",
        style
    )

    frame = Frame(50, height - 700, 500, 200)  # x, y, largura, altura
    frame.addFromList([paragrafo], c)
    
# 1 Deficiência ou Necessidade Especial
    paragrafo = Paragraph(
        "1. Seu filho(a) possui alguma deficiência que devemos conhecer para acolhê-lo(a) "
        "da melhor forma?",
        style
    )
    frame = Frame(50, height - 750, 500, 200)
    frame.addFromList([paragrafo], c)
    if ficha.possui_deficiencia:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_deficiencia}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)

# 2 Transtornos
    paragrafo = Paragraph(
        "2. Há algum transtorno ou diagnóstico que ajude nossa pastoral a compreender melhor as necessidades da criança?",
        style
    )
    frame = Frame(50, height - 795, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.possui_transtorno:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_transtorno}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 825, 500, 200)
    frame.addFromList([paragrafo], c)
    
# 3 Medicamentos
    paragrafo = Paragraph(
        "3. Seu filho(a) faz uso de algum medicamento contínuo?",
        style
    )
    frame = Frame(50, height - 840, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.medicamento_uso_continuo:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_medicamento}. Horário: {ficha.medicamento_horario}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 855, 500, 200)
    frame.addFromList([paragrafo], c)   
    
# 4 Acompanhamento Psicológico
    paragrafo = Paragraph(
        "4. A criança faz acompanhamento psicológico, psiquiátrico ou terapêutico?",
        style
    )
    frame = Frame(50, height - 870, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.acompanhamento_psicologico:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_acompanhamento}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 885, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            "Agradecemos de coração pela confiança. Nosso desejo é caminhar juntos para que o(a) jovem viva cada momento com alegria, segurança e acolhimento.",
        style
    )
    frame = Frame(50, height - 900, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> Assinatura do Responsável",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    
    
    c.showPage()  # Página 2
    

    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"TERMO DE CONSENTIMENTO PARA TRATAMENTO DE DADOS ")
    c.drawString(120, height - 180, f"PESSOAIS SENSÍVEIS DE CRIANÇAS E ADOLESCENTES")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome_responsavel}, CPF {ficha.cpf_responsavel} na qualidade de responsável legal pelo(a) menor "
        f"{ficha.nome}, por meio deste instrumento, <b>manifesto meu consentimento livre, informado e inequívoco para o tratamento dos dados pessoais do(a) referido"
        "menor</b>, nos seguintes termos, em conformidade com a Lei no 13.709/2018:"
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"1) Autorizo o tratamento a ser realizado pela PARÓQUIA NOSSA SENHORA APARECIDA, pessoa jurídica "
        f"de direito privado, inscrita no CNPJ no 44.802.999/0011-30, situada na Rua dois, no 349, Bairro Aparecida,"
        f"Rio Claro/SP, CEP 13500-270, e e-mail: pnsarc@hotmail.com, doravante denominada CONTROLADORA."
        ,style
    )
    frame = Frame(50, height - 460, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"2) Os dados pessoais do(a) menor serão utilizados exclusivamente para:<br/>"
        f"    I) Inscrição e organização da Catequese;<br/>"
        f"    II) Emitir certificados de conclusão dos sacramentos;<br/>"
        f"    III) Permitir contato com pais e/ou responsáveis.<br/>"
        ,style
    )
    frame = Frame(50, height - 520, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"3) Autorizo o tratamento apenas dos seguintes dados pessoais: nome completo; data de nascimento;"
        f"naturalidade; documento de identificação (RG, CPF, certidão de nascimento ou equivalente); endereço completo; telefone de contato."
        f"Não haverá utilização dos dados para fins comerciais, divulgação indevida ou compartilhamento com terceiros estranhos às atividades religiosas."
        ,style
    )
    frame = Frame(50, height - 590, 500, 200)
    frame.addFromList([paragrafo], c)
    

    paragrafo = Paragraph(
        f"4) Declaro estar ciente de que o armazenamento ocorrerá:"
        f"I) Em fichas físicas e/ou sistemas informatizados da Paróquia;<br/>"
        f"II) Que o acesso será restrito a pessoas autorizadas (secretaria, coordenação da catequese, catequistas e o Pároco, quando necessário);<br/>"
        f"III) Que a CONTROLADORA adotará medidas técnicas e administrativas adequadas para proteger os dados contra acessos não autorizados, perda, divulgação indevida ou qualquer forma de tratamento inadequado.<br/>"
        ,style
    )
    frame = Frame(50, height - 670, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"5) Os dados serão mantidos enquanto o(a) menor estiver matriculado(a) na Catequese, pelo prazo máximo de 3 (três) anos, quando concluída as etapas sacramentais, respeitando-se a necessidade de registros paroquiais.<br/>"
        f"6) Estou ciente de que posso, a qualquer momento solicitar acesso aos dados pessoais do(a) menor; solicitar correção de dados incompletos, inexatos ou desatualizados; requerer a eliminação de dados desnecessários, excessivos ou tratados em desconformidade com a lei.<br/>"
        f"7) Declaro, ainda, estar ciente de que posso revogar este consentimento a qualquer momento, mediante solicitação formal e por escrito à Secretaria Paroquial.<br/>"
        ,style
    )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)
   
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome_responsavel} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.showPage()  # Página 3 
    
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"AUTORIZAÇÃO PARA USO DE IMAGEM MENORES DE 18 ANOS")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome_responsavel}, CPF {ficha.cpf_responsavel} na qualidade de responsável legal pelo(a) menor "
        f"{ficha.nome}, autorizo, de forma livre, expressa e informada, a Paróquia Nossa Senhora Aparecida, inscrita "
        f"no CNPJ sob o nº 44.802.999/0011-30, a utilizar a imagem, nome e voz do(a) menor, captados em fotografias e/ou vídeos durante atividades e eventos da paróquia, para fins de divulgação em meios impressos, digitais e redes sociais da paróquia, sem qualquer ônus."
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Declaro estar ciente de que a utilização da imagem do(a) menor será feita de acordo com a Lei Geral de Proteção de Dados (Lei nº 13.709/2018 - LGPD), e que posso, a qualquer momento, revogar esta autorização mediante solicitação por escrito à paróquia. "
        ,style
    )
    frame = Frame(50, height - 480, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Estou ciente de que não tenho direito a qualquer remuneração pelo uso da imagem, nome e voz do(a) menor nos termos acima mencionados, e que a presente autorização é concedida por prazo indeterminado, podendo ser revogada a qualquer momento, mediante comunicação por escrito. "
        ,style
    )
    frame = Frame(50, height - 530, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"Por fim, declaro que a presente autorização foi feita de forma livre, sem qualquer coação, e que fui devidamente informado(a) sobre o tratamento dos dados pessoais do(a) menor pela paróquia. "
        ,style
    )
    frame = Frame(50, height - 580, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome_responsavel} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 660, 500, 200)
    frame.addFromList([paragrafo], c)
    
    
    c.save()

    return filename

def gerar_ficha_crisma_maior(ficha):
    img_path = os.path.join(settings.BASE_DIR, 'static', 'pdf', 'cabecalho.png')

    filename = os.path.join(settings.MEDIA_ROOT, f"ficha_crisma_{ficha.id}.pdf")

    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4

    # Inserindo imagem no topo
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 200, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica", 16)
    c.drawString(190, height - 220, f"INSCRIÇÃO PARA CRISMA")
    
    # Nome
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 250, f"Nome:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 250, f"{ficha.nome}")

    # Sexo
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 265, f"Sexo:")
    c.setFont("Helvetica", 11)
    if ficha.sexo == 'M':
        c.drawString(150, height - 265, f"Masculino")
    else:
        c.drawString(150, height - 265, f"Feminino")
    
    # Data de Nascimento
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 280, f"Data Nascimento:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 280, f"{ficha.data_nascimento.strftime("%d/%m/%Y")}")
    
    # Naturalidade
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 295, f"Naturalidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 295, f"{ficha.naturalidade}")
    
    # Nome dos Pais
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 310, f"Pai:")
    c.setFont("Helvetica", 11)
    if ficha.nome_pai:
        c.drawString(150, height - 310, f"{ficha.nome_pai}  -  {ficha.celular_pai}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 325, f"Mãe:")
    c.setFont("Helvetica", 11)
    if ficha.nome_mae:
        c.drawString(150, height - 325, f"{ficha.nome_mae}  -  {ficha.celular_mae}")

    # Endereço
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 340, f"Endereço:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 340, f"{ficha.endereco}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 355, f"Cidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 355, f"{ficha.cidade}  -  {ficha.uf}")
    
    # Batizado
    c.setFont("Helvetica-Bold", 11)
    if ficha.batizado:
        c.drawString(50, height - 375, f"Batizado")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 390, f"Data: {ficha.batizado_data.strftime("%d/%m/%Y") if ficha.batizado_data else ''}")
        c.drawString(50, height - 405, f"Diocese: {ficha.batizado_diocese}   Paróquia: {ficha.batizado_paroquia}")
        c.drawString(50, height - 420, f"Celebrante: {ficha.batizado_celebrante} ")
        
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 375, f"Não Batizado")
        c.setFillColor(colors.black)
    
    
    c.setFont("Helvetica", 11)
    c.drawString(50, height - 440, f"Horário:")
    c.setFillColor(colors.blue)
    c.drawString(150, height - 440, f"{ficha.get_horario_display()} ")
    c.setFillColor(colors.black)
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

# Cuidado e Acolhimento da Criança
    c.setFont("Helvetica-Bold", 16)
    c.drawString(170, height - 480, "Cuidado e Acolhimento")   
    c.setFont("Helvetica", 11)
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        "Para que possamos receber você com todo carinho, atenção e segurança, "
        "pedimos que compartilhe conosco algumas informações importantes. "
        "Tudo será tratado com sigilo e usado apenas para ajudar no seu bem-estar "
        "durante as atividades.",
        style
    )

    frame = Frame(50, height - 700, 500, 200)  # x, y, largura, altura
    frame.addFromList([paragrafo], c)
    
# 1 Deficiência ou Necessidade Especial
    paragrafo = Paragraph(
        "1. Você possui alguma deficiência que devemos conhecer para acolhê-lo(a) "
        "da melhor forma?",
        style
    )
    frame = Frame(50, height - 750, 500, 200)
    frame.addFromList([paragrafo], c)
    if ficha.possui_deficiencia:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_deficiencia}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)

# 2 Transtornos
    paragrafo = Paragraph(
        "2. Há algum transtorno ou diagnóstico que ajude nossa pastoral a compreender melhor as necessidades da criança?",
        style
    )
    frame = Frame(50, height - 795, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.possui_transtorno:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_transtorno}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 825, 500, 200)
    frame.addFromList([paragrafo], c)
    
# 3 Medicamentos
    paragrafo = Paragraph(
        "3. Você faz uso de algum medicamento contínuo?",
        style
    )
    frame = Frame(50, height - 840, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.medicamento_uso_continuo:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_medicamento}. Horário: {ficha.medicamento_horario}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 855, 500, 200)
    frame.addFromList([paragrafo], c)   
    
# 4 Acompanhamento Psicológico
    paragrafo = Paragraph(
        "4. Você faz acompanhamento psicológico, psiquiátrico ou terapêutico?",
        style
    )
    frame = Frame(50, height - 870, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.acompanhamento_psicologico:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_acompanhamento}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 885, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            "Agradecemos de coração pela confiança. Nosso desejo é caminhar juntos para que você viva cada momento com alegria, segurança e acolhimento.",
        style
    )
    frame = Frame(50, height - 900, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> Assinatura do Responsável",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    
    
    c.showPage()  # Página 2
    

    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"TERMO DE CONSENTIMENTO PARA TRATAMENTO ")
    c.drawString(150, height - 180, f"DE DADOS PESSOAIS SENSÍVEIS")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome}, CPF {ficha.cpf_responsavel}, "
        f"por meio deste instrumento, <b>manifesto meu consentimento livre, informado e inequívoco para o tratamento dos meus dados pessoais</b>,"
        "nos seguintes termos, em conformidade com a Lei no 13.709/2018:"
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"1) Autorizo o tratamento a ser realizado pela PARÓQUIA NOSSA SENHORA APARECIDA, pessoa jurídica "
        f"de direito privado, inscrita no CNPJ no 44.802.999/0011-30, situada na Rua dois, no 349, Bairro Aparecida,"
        f"Rio Claro/SP, CEP 13500-270, e e-mail: pnsarc@hotmail.com, doravante denominada CONTROLADORA."
        ,style
    )
    frame = Frame(50, height - 460, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"2) Os meus dados pessoais serão utilizados exclusivamente para:<br/>"
        f"    I) Inscrição e organização da Catequese;<br/>"
        f"    II) Emitir certificados de conclusão dos sacramentos;<br/>"
        f"    III) Permitir contato comigo e/ou com pais e/ou responsáveis.<br/>"
        ,style
    )
    frame = Frame(50, height - 520, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"3) Autorizo o tratamento apenas dos seguintes dados pessoais: nome completo; data de nascimento;"
        f"naturalidade; documento de identificação (RG, CPF, certidão de nascimento ou equivalente); endereço completo; telefone de contato."
        f"Não haverá utilização dos dados para fins comerciais, divulgação indevida ou compartilhamento com terceiros estranhos às atividades religiosas."
        ,style
    )
    frame = Frame(50, height - 590, 500, 200)
    frame.addFromList([paragrafo], c)
    

    paragrafo = Paragraph(
        f"4) Declaro estar ciente de que o armazenamento ocorrerá:"
        f"I) Em fichas físicas e/ou sistemas informatizados da Paróquia;<br/>"
        f"II) Que o acesso será restrito a pessoas autorizadas (secretaria, coordenação da catequese, catequistas e o Pároco, quando necessário);<br/>"
        f"III) Que a CONTROLADORA adotará medidas técnicas e administrativas adequadas para proteger os dados contra acessos não autorizados, perda, divulgação indevida ou qualquer forma de tratamento inadequado.<br/>"
        ,style
    )
    frame = Frame(50, height - 670, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"5) Os dados serão mantidos enquanto eu estiver matriculado(a) na Crisma, pelo prazo máximo de 3 (três) anos, quando concluída as etapas sacramentais, respeitando-se a necessidade de registros paroquiais.<br/>"
        f"6) Estou ciente de que posso, a qualquer momento solicitar acesso aos meus dados pessoais; solicitar correção de dados incompletos, inexatos ou desatualizados; requerer a eliminação de dados desnecessários, excessivos ou tratados em desconformidade com a lei.<br/>"
        f"7) Declaro, ainda, estar ciente de que posso revogar este consentimento a qualquer momento, mediante solicitação formal e por escrito à Secretaria Paroquial.<br/>"
        ,style
    )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)
   
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.showPage()  # Página 3 
    
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"AUTORIZAÇÃO PARA USO DE IMAGEM ")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome}, CPF {ficha.cpf_responsavel}, "
        f"autorizo, de forma livre, expressa e informada, a Paróquia Nossa Senhora Aparecida, inscrita "
        f"no CNPJ sob o nº 44.802.999/0011-30, a utilizar a minha imagem, nome e voz, captados em fotografias e/ou vídeos durante atividades e eventos da paróquia, para fins de divulgação em meios impressos, digitais e redes sociais da paróquia, sem qualquer ônus."
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Declaro estar ciente de que a utilização da minha imagem será feita de acordo com a Lei Geral de Proteção de Dados (Lei nº 13.709/2018 - LGPD), e que posso, a qualquer momento, revogar esta autorização mediante solicitação por escrito à paróquia. "
        ,style
    )
    frame = Frame(50, height - 480, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Estou ciente de que não tenho direito a qualquer remuneração pelo uso da minha imagem, nome e voz nos termos acima mencionados, e que a presente autorização é concedida por prazo indeterminado, podendo ser revogada a qualquer momento, mediante comunicação por escrito. "
        ,style
    )
    frame = Frame(50, height - 530, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"Por fim, declaro que a presente autorização foi feita de forma livre, sem qualquer coação, e que fui devidamente informado(a) sobre o tratamento dos meus dados pessoais pela paróquia. "
        ,style
    )
    frame = Frame(50, height - 580, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome_responsavel} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 660, 500, 200)
    frame.addFromList([paragrafo], c)
    
    
    c.save()

    return filename

def gerar_ficha_perseveranca_mej(ficha):
    anos = ((date.today() - ficha.data_nascimento)/365.2425).days
    if anos < 18:
        return gerar_ficha_perseveranca_mej_menor_idade(ficha)
    if anos >= 18:
        return gerar_ficha_perseveranca_mej_maior_idade(ficha)

def gerar_ficha_perseveranca_mej_menor_idade(ficha):
    img_path = os.path.join(settings.BASE_DIR, 'static', 'pdf', 'cabecalho.png')

    filename = os.path.join(settings.MEDIA_ROOT, f"ficha_perseveranca_mej_{ficha.id}.pdf")

    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4

    # Inserindo imagem no topo
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 200, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica", 16)
    c.drawString(130, height - 220, f"INSCRIÇÃO PARA PERSEVERANCA / MEJ")
    
    # Nome
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 250, f"Nome:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 250, f"{ficha.nome}")

    # Sexo
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 265, f"Sexo:")
    c.setFont("Helvetica", 11)
    if ficha.sexo == 'M':
        c.drawString(150, height - 265, f"Masculino")
    else:
        c.drawString(150, height - 265, f"Feminino")
    
    # Data de Nascimento
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 280, f"Data Nascimento:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 280, f"{ficha.data_nascimento.strftime("%d/%m/%Y")}")
    
    # Naturalidade
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 295, f"Naturalidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 295, f"{ficha.naturalidade}")
    
    # Nome dos Pais
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 310, f"Pai:")
    c.setFont("Helvetica", 11)
    if ficha.nome_pai:
        c.drawString(150, height - 310, f"{ficha.nome_pai}  -  {ficha.celular_pai}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 325, f"Mãe:")
    c.setFont("Helvetica", 11)
    if ficha.nome_mae:
        c.drawString(150, height - 325, f"{ficha.nome_mae}  -  {ficha.celular_mae}")

    # Endereço
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 340, f"Endereço:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 340, f"{ficha.endereco}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 355, f"Cidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 355, f"{ficha.cidade}  -  {ficha.uf}")
    
    # Batizado
    c.setFont("Helvetica", 11)
    if ficha.batizado:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 375, f"Batizado na Data: {ficha.batizado_data.strftime("%d/%m/%Y") if ficha.batizado_data else ''}")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 385, f"Diocese: {ficha.batizado_diocese} Celebrante: {ficha.batizado_celebrante}  Paróquia: {ficha.batizado_paroquia}")
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 375, f"Não Batizado")
        c.setFillColor(colors.black)

# Primeira Eucaristia
    if ficha.primeira_eucaristia:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 400, f"Primeira Eucaristia - Data: {ficha.primeira_eucaristia_data.strftime("%d/%m/%Y") if ficha.primeira_eucaristia_data else ''}")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 410, f"Diocese: {ficha.primeira_eucaristia_diocese}   Paróquia: {ficha.primeira_eucaristia_paroquia} - Celebrante: {ficha.primeira_eucaristia_celebrante} ")
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 400, f"Não Fez Primeira Eucaristia")
        c.setFillColor(colors.black)
    
    c.setFont("Helvetica", 11)
    c.drawString(50, height - 440, f"Horário:")
    if ficha.horario == '1':
        c.setFillColor(colors.blue)
        c.drawString(150, height - 440, f"{ficha.get_horario_display()}")
    else:
        c.setFillColor(colors.red)
        c.drawString(150, height - 440, f"{ficha.get_horario_display()}")
    c.setFillColor(colors.black)
    
# Cuidado e Acolhimento da Criança
    c.setFont("Helvetica-Bold", 16)
    c.drawString(170, height - 480, "Cuidado e Acolhimento da Criança")   
    c.setFont("Helvetica", 11)
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        "Para que possamos receber seu filho(a) com todo carinho, atenção e segurança, "
        "pedimos que compartilhe conosco algumas informações importantes. "
        "Tudo será tratado com sigilo e usado apenas para ajudar no bem-estar "
        "da criança durante as atividades.",
        style
    )

    frame = Frame(50, height - 700, 500, 200)  # x, y, largura, altura
    frame.addFromList([paragrafo], c)
    
# 1 Deficiência ou Necessidade Especial
    paragrafo = Paragraph(
        "1. Seu filho(a) possui alguma deficiência que devemos conhecer para acolhê-lo(a) "
        "da melhor forma?",
        style
    )
    frame = Frame(50, height - 750, 500, 200)
    frame.addFromList([paragrafo], c)
    if ficha.possui_deficiencia:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_deficiencia}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)

# 2 Transtornos
    paragrafo = Paragraph(
        "2. Há algum transtorno ou diagnóstico que ajude nossa pastoral a compreender melhor as necessidades da criança?",
        style
    )
    frame = Frame(50, height - 795, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.possui_transtorno:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_transtorno}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 825, 500, 200)
    frame.addFromList([paragrafo], c)
    
# 3 Medicamentos
    paragrafo = Paragraph(
        "3. Seu filho(a) faz uso de algum medicamento contínuo?",
        style
    )
    frame = Frame(50, height - 840, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.medicamento_uso_continuo:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_medicamento}. Horário: {ficha.medicamento_horario}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 855, 500, 200)
    frame.addFromList([paragrafo], c)   
    
# 4 Acompanhamento Psicológico
    paragrafo = Paragraph(
        "4. A criança faz acompanhamento psicológico, psiquiátrico ou terapêutico?",
        style
    )
    frame = Frame(50, height - 870, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.acompanhamento_psicologico:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_acompanhamento}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 885, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            "Agradecemos de coração pela confiança. Nosso desejo é caminhar juntos para que a criança viva cada momento com alegria, segurança e acolhimento.",
        style
    )
    frame = Frame(50, height - 900, 500, 200)
    frame.addFromList([paragrafo], c)
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> Assinatura do Responsável",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    

    c.showPage()  # Página 2

    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"TERMO DE CONSENTIMENTO PARA TRATAMENTO DE DADOS ")
    c.drawString(120, height - 180, f"PESSOAIS SENSÍVEIS DE CRIANÇAS E ADOLESCENTES")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome_responsavel}, CPF {ficha.cpf_responsavel} na qualidade de responsável legal pelo(a) menor "
        f"{ficha.nome}, por meio deste instrumento, <b>manifesto meu consentimento livre, informado e inequívoco para o tratamento dos dados pessoais do(a) referido"
        "menor</b>, nos seguintes termos, em conformidade com a Lei no 13.709/2018:"
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"1) Autorizo o tratamento a ser realizado pela PARÓQUIA NOSSA SENHORA APARECIDA, pessoa jurídica "
        f"de direito privado, inscrita no CNPJ no 44.802.999/0011-30, situada na Rua dois, no 349, Bairro Aparecida,"
        f"Rio Claro/SP, CEP 13500-270, e e-mail: pnsarc@hotmail.com, doravante denominada CONTROLADORA."
        ,style
    )
    frame = Frame(50, height - 460, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"2) Os dados pessoais do(a) menor serão utilizados exclusivamente para:<br/>"
        f"    I) Inscrição e organização da Catequese;<br/>"
        f"    II) Emitir certificados de conclusão dos sacramentos;<br/>"
        f"    III) Permitir contato com pais e/ou responsáveis.<br/>"
        ,style
    )
    frame = Frame(50, height - 520, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"3) Autorizo o tratamento apenas dos seguintes dados pessoais: nome completo; data de nascimento;"
        f"naturalidade; documento de identificação (RG, CPF, certidão de nascimento ou equivalente); endereço completo; telefone de contato."
        f"Não haverá utilização dos dados para fins comerciais, divulgação indevida ou compartilhamento com terceiros estranhos às atividades religiosas."
        ,style
    )
    frame = Frame(50, height - 590, 500, 200)
    frame.addFromList([paragrafo], c)
    

    paragrafo = Paragraph(
        f"4) Declaro estar ciente de que o armazenamento ocorrerá:"
        f"I) Em fichas físicas e/ou sistemas informatizados da Paróquia;<br/>"
        f"II) Que o acesso será restrito a pessoas autorizadas (secretaria, coordenação da catequese, catequistas e o Pároco, quando necessário);<br/>"
        f"III) Que a CONTROLADORA adotará medidas técnicas e administrativas adequadas para proteger os dados contra acessos não autorizados, perda, divulgação indevida ou qualquer forma de tratamento inadequado.<br/>"
        ,style
    )
    frame = Frame(50, height - 670, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"5) Os dados serão mantidos enquanto o(a) menor estiver matriculado(a) na Catequese, pelo prazo máximo de 3 (três) anos, quando concluída as etapas sacramentais, respeitando-se a necessidade de registros paroquiais.<br/>"
        f"6) Estou ciente de que posso, a qualquer momento solicitar acesso aos dados pessoais do(a) menor; solicitar correção de dados incompletos, inexatos ou desatualizados; requerer a eliminação de dados desnecessários, excessivos ou tratados em desconformidade com a lei.<br/>"
        f"7) Declaro, ainda, estar ciente de que posso revogar este consentimento a qualquer momento, mediante solicitação formal e por escrito à Secretaria Paroquial.<br/>"
        ,style
    )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)
   
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome_responsavel} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.showPage()  # Página 3 
    
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"AUTORIZAÇÃO PARA USO DE IMAGEM MENORES DE 18 ANOS")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome_responsavel}, CPF {ficha.cpf_responsavel} na qualidade de responsável legal pelo(a) menor "
        f"{ficha.nome}, autorizo, de forma livre, expressa e informada, a Paróquia Nossa Senhora Aparecida, inscrita "
        f"no CNPJ sob o nº 44.802.999/0011-30, a utilizar a imagem, nome e voz do(a) menor, captados em fotografias e/ou vídeos durante atividades e eventos da paróquia, para fins de divulgação em meios impressos, digitais e redes sociais da paróquia, sem qualquer ônus."
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Declaro estar ciente de que a utilização da imagem do(a) menor será feita de acordo com a Lei Geral de Proteção de Dados (Lei nº 13.709/2018 - LGPD), e que posso, a qualquer momento, revogar esta autorização mediante solicitação por escrito à paróquia. "
        ,style
    )
    frame = Frame(50, height - 480, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Estou ciente de que não tenho direito a qualquer remuneração pelo uso da imagem, nome e voz do(a) menor nos termos acima mencionados, e que a presente autorização é concedida por prazo indeterminado, podendo ser revogada a qualquer momento, mediante comunicação por escrito. "
        ,style
    )
    frame = Frame(50, height - 530, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"Por fim, declaro que a presente autorização foi feita de forma livre, sem qualquer coação, e que fui devidamente informado(a) sobre o tratamento dos dados pessoais do(a) menor pela paróquia. "
        ,style
    )
    frame = Frame(50, height - 580, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome_responsavel} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 660, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.save()

    return filename

def gerar_ficha_perseveranca_mej_maior_idade(ficha):
    img_path = os.path.join(settings.BASE_DIR, 'static', 'pdf', 'cabecalho.png')

    filename = os.path.join(settings.MEDIA_ROOT, f"ficha_perseveranca_mej_{ficha.id}.pdf")

    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4

    # Inserindo imagem no topo
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 200, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica", 16)
    c.drawString(130, height - 220, f"INSCRIÇÃO PARA PERSEVERANCA / MEJ")
    
    # Nome
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 250, f"Nome:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 250, f"{ficha.nome}")

    # Sexo
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 265, f"Sexo:")
    c.setFont("Helvetica", 11)
    if ficha.sexo == 'M':
        c.drawString(150, height - 265, f"Masculino")
    else:
        c.drawString(150, height - 265, f"Feminino")
    
    # Data de Nascimento
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 280, f"Data Nascimento:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 280, f"{ficha.data_nascimento.strftime("%d/%m/%Y")}")
    
    # Naturalidade
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 295, f"Naturalidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 295, f"{ficha.naturalidade}")
    
    # Nome dos Pais
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 310, f"Pai:")
    c.setFont("Helvetica", 11)
    if ficha.nome_pai:
        c.drawString(150, height - 310, f"{ficha.nome_pai}  -  {ficha.celular_pai}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 325, f"Mãe:")
    c.setFont("Helvetica", 11)
    if ficha.nome_mae:
        c.drawString(150, height - 325, f"{ficha.nome_mae}  -  {ficha.celular_mae}")

    # Endereço
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 340, f"Endereço:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 340, f"{ficha.endereco}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 355, f"Cidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 355, f"{ficha.cidade}  -  {ficha.uf}")
    
    # Batizado
    c.setFont("Helvetica", 11)
    if ficha.batizado:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 375, f"Batizado na Data: {ficha.batizado_data.strftime("%d/%m/%Y") if ficha.batizado_data else ''}")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 385, f"Diocese: {ficha.batizado_diocese} Celebrante: {ficha.batizado_celebrante}  Paróquia: {ficha.batizado_paroquia}")
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 375, f"Não Batizado")
        c.setFillColor(colors.black)

# Primeira Eucaristia
    if ficha.primeira_eucaristia:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 400, f"Primeira Eucaristia - Data: {ficha.primeira_eucaristia_data.strftime("%d/%m/%Y") if ficha.primeira_eucaristia_data else ''}")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 410, f"Diocese: {ficha.primeira_eucaristia_diocese}   Paróquia: {ficha.primeira_eucaristia_paroquia} - Celebrante: {ficha.primeira_eucaristia_celebrante} ")
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 400, f"Não Fez Primeira Eucaristia")
        c.setFillColor(colors.black)
    
    c.setFont("Helvetica", 11)
    c.drawString(50, height - 440, f"Horário:")
    if ficha.horario == '1':
        c.setFillColor(colors.blue)
        c.drawString(150, height - 440, f"{ficha.get_horario_display()}")
    else:
        c.setFillColor(colors.red)
        c.drawString(150, height - 440, f"{ficha.get_horario_display()}")
    c.setFillColor(colors.black)
    
# Cuidado e Acolhimento da Criança
    c.setFont("Helvetica-Bold", 16)
    c.drawString(200, height - 480, "Cuidado e Acolhimento")   
    c.setFont("Helvetica", 11)
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        "Para que possamos lhe receber com todo carinho, atenção e segurança, "
        "pedimos que compartilhe conosco algumas informações importantes. "
        "Tudo será tratado com sigilo e usado apenas para ajudar no seu bem-estar "
        " durante as atividades.",
        style
    )

    frame = Frame(50, height - 700, 500, 200)  # x, y, largura, altura
    frame.addFromList([paragrafo], c)
    
# 1 Deficiência ou Necessidade Especial
    paragrafo = Paragraph(
        "1. Você possui alguma deficiência que devemos conhecer para acolhê-lo(a) "
        "da melhor forma?",
        style
    )
    frame = Frame(50, height - 750, 500, 200)
    frame.addFromList([paragrafo], c)
    if ficha.possui_deficiencia:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_deficiencia}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)

# 2 Transtornos
    paragrafo = Paragraph(
        "2. Há algum transtorno ou diagnóstico que ajude nossa pastoral a compreender melhor as suas necessidades ?",
        style
    )
    frame = Frame(50, height - 795, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.possui_transtorno:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_transtorno}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 825, 500, 200)
    frame.addFromList([paragrafo], c)
    
# 3 Medicamentos
    paragrafo = Paragraph(
        "3. Você faz uso de algum medicamento contínuo?",
        style
    )
    frame = Frame(50, height - 840, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.medicamento_uso_continuo:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_medicamento}. Horário: {ficha.medicamento_horario}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 855, 500, 200)
    frame.addFromList([paragrafo], c)   
    
# 4 Acompanhamento Psicológico
    paragrafo = Paragraph(
        "4. Você faz acompanhamento psicológico, psiquiátrico ou terapêutico?",
        style
    )
    frame = Frame(50, height - 870, 500, 200)
    frame.addFromList([paragrafo], c)    
    if ficha.acompanhamento_psicologico:
        paragrafo = Paragraph(
            f"Sim. Descrição: {ficha.descricao_acompanhamento}",
            style
        )
    else:
        paragrafo = Paragraph(
            "Não.",
            style
        )
    frame = Frame(50, height - 885, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            "Agradecemos de coração pela confiança. Nosso desejo é caminhar juntos para que você viva cada momento com alegria, segurança e acolhimento.",
        style
    )
    frame = Frame(50, height - 900, 500, 200)
    frame.addFromList([paragrafo], c)
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> Assinatura do Responsável",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    

    c.showPage()  # Página 2

    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"TERMO DE CONSENTIMENTO PARA TRATAMENTO ")
    c.drawString(160, height - 180, f"DE DADOS PESSOAIS SENSÍVEIS")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome}, CPF {ficha.cpf_responsavel} por meio deste instrumento, <b>manifesto meu consentimento livre, informado e inequívoco para o tratamento dos meus dados pessoais"
        "</b>, nos seguintes termos, em conformidade com a Lei no 13.709/2018:"
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"1) Autorizo o tratamento a ser realizado pela PARÓQUIA NOSSA SENHORA APARECIDA, pessoa jurídica "
        f"de direito privado, inscrita no CNPJ no 44.802.999/0011-30, situada na Rua dois, no 349, Bairro Aparecida,"
        f"Rio Claro/SP, CEP 13500-270, e e-mail: pnsarc@hotmail.com, doravante denominada CONTROLADORA."
        ,style
    )
    frame = Frame(50, height - 460, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"2) Os meus dados pessoais serão utilizados exclusivamente para:<br/>"
        f"    I) Inscrição e organização da Catequese / Perseverança / MEJ;<br/>"
        f"    II) Emitir certificados de conclusão dos sacramentos;<br/>"
        f"    III) Permitir contato com pais e/ou responsáveis.<br/>"
        ,style
    )
    frame = Frame(50, height - 520, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"3) Autorizo o tratamento apenas dos seguintes dados pessoais: nome completo; data de nascimento;"
        f"naturalidade; documento de identificação (RG, CPF, certidão de nascimento ou equivalente); endereço completo; telefone de contato."
        f"Não haverá utilização dos dados para fins comerciais, divulgação indevida ou compartilhamento com terceiros estranhos às atividades religiosas."
        ,style
    )
    frame = Frame(50, height - 590, 500, 200)
    frame.addFromList([paragrafo], c)
    

    paragrafo = Paragraph(
        f"4) Declaro estar ciente de que o armazenamento ocorrerá:"
        f"I) Em fichas físicas e/ou sistemas informatizados da Paróquia;<br/>"
        f"II) Que o acesso será restrito a pessoas autorizadas (secretaria, coordenação da catequese, catequistas e o Pároco, quando necessário);<br/>"
        f"III) Que a CONTROLADORA adotará medidas técnicas e administrativas adequadas para proteger os dados contra acessos não autorizados, perda, divulgação indevida ou qualquer forma de tratamento inadequado.<br/>"
        ,style
    )
    frame = Frame(50, height - 670, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"5) Os dados serão mantidos enquanto eu estiver matriculado(a) na Perseverança/MEJ, pelo prazo máximo de 3 (três) anos, quando concluída as etapas sacramentais, respeitando-se a necessidade de registros paroquiais.<br/>"
        f"6) Estou ciente de que posso, a qualquer momento solicitar acesso aos meus dados pessoais; solicitar correção de dados incompletos, inexatos ou desatualizados; requerer a eliminação de dados desnecessários, excessivos ou tratados em desconformidade com a lei.<br/>"
        f"7) Declaro, ainda, estar ciente de que posso revogar este consentimento a qualquer momento, mediante solicitação formal e por escrito à Secretaria Paroquial.<br/>"
        ,style
    )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)
   
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.showPage()  # Página 3 
    
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"AUTORIZAÇÃO PARA USO DE IMAGEM ")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome}, CPF {ficha.cpf_responsavel}, "
        f"autorizo, de forma livre, expressa e informada, a Paróquia Nossa Senhora Aparecida, inscrita "
        f"no CNPJ sob o nº 44.802.999/0011-30, a utilizar a minha imagem, nome e voz, captados em fotografias e/ou vídeos durante atividades e eventos da paróquia, para fins de divulgação em meios impressos, digitais e redes sociais da paróquia, sem qualquer ônus."
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Declaro estar ciente de que a utilização da minha imagem será feita de acordo com a Lei Geral de Proteção de Dados (Lei nº 13.709/2018 - LGPD), e que posso, a qualquer momento, revogar esta autorização mediante solicitação por escrito à paróquia. "
        ,style
    )
    frame = Frame(50, height - 480, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Estou ciente de que não tenho direito a qualquer remuneração pelo uso da minha imagem, nome e voz nos termos acima mencionados, e que a presente autorização é concedida por prazo indeterminado, podendo ser revogada a qualquer momento, mediante comunicação por escrito. "
        ,style
    )
    frame = Frame(50, height - 530, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"Por fim, declaro que a presente autorização foi feita de forma livre, sem qualquer coação, e que fui devidamente informado(a) sobre o tratamento dos meus dados pessoais pela paróquia. "
        ,style
    )
    frame = Frame(50, height - 580, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome} - CPF: {ficha.cpf_responsavel}",
        style
    )
    frame = Frame(250, height - 660, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.save()
    return filename

def gerar_ficha_catequese_adulto(ficha):
    img_path = os.path.join(settings.BASE_DIR, 'static', 'pdf', 'cabecalho.png')

    filename = os.path.join(settings.MEDIA_ROOT, f"ficha_catequese_adulto_{ficha.id}.pdf")

    c = canvas.Canvas(filename, pagesize=A4)
    width, height = A4

    # Inserindo imagem no topo
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 200, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica", 16)
    c.drawString(130, height - 220, f"INSCRIÇÃO PARA CATEQUESE ADULTO")
    
    # Nome
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 250, f"Nome:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 250, f"{ficha.nome}")

    # Sexo
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 265, f"Sexo:")
    c.setFont("Helvetica", 11)
    if ficha.sexo == 'M':
        c.drawString(150, height - 265, f"Masculino")
    else:
        c.drawString(150, height - 265, f"Feminino")
    
    # Data de Nascimento
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 280, f"Data Nascimento:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 280, f"{ficha.data_nascimento.strftime("%d/%m/%Y")}")
    
    # Naturalidade
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 295, f"Naturalidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 295, f"{ficha.naturalidade}")
    
    # Nome dos Pais
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 310, f"Pai:")
    c.setFont("Helvetica", 11)
    if ficha.nome_pai:
        c.drawString(150, height - 310, f"{ficha.nome_pai}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 325, f"Mãe:")
    c.setFont("Helvetica", 11)
    if ficha.nome_mae:
        c.drawString(150, height - 325, f"{ficha.nome_mae}")

    # Endereço
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 340, f"Endereço:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 340, f"{ficha.endereco}")
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 355, f"Cidade:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 355, f"{ficha.cidade}  -  {ficha.uf}")
    
    c.setFont("Helvetica-Bold", 11)
    c.drawString(50, height - 370, f"Estado Civil:")
    c.setFont("Helvetica", 11)
    c.drawString(150, height - 370, f"{ficha.estado_civil}")
    
    # Batizado
    c.setFont("Helvetica", 11)
    if ficha.batizado:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 400, f"Batizado na Data: {ficha.batizado_data.strftime("%d/%m/%Y") if ficha.batizado_data else ''}")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 415, f"Diocese: {ficha.batizado_diocese}")
        c.drawString(50, height - 430, f"Paróquia: {ficha.batizado_paroquia}")
        c.drawString(50, height - 445, f"Celebrante: {ficha.batizado_celebrante}")
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 400, f"Não Batizado")
        c.setFillColor(colors.black)

# Primeira Eucaristia
    if ficha.primeira_eucaristia:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 500, f"Primeira Eucaristia - Data: {ficha.primeira_eucaristia_data.strftime("%d/%m/%Y") if ficha.primeira_eucaristia_data else ''}")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 515, f"Diocese: {ficha.primeira_eucaristia_diocese}")
        c.drawString(50, height - 530, f"Paróquia: {ficha.primeira_eucaristia_paroquia}")
        c.drawString(50, height - 545, f"Celebrante: {ficha.primeira_eucaristia_celebrante} ")
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 500, f"Não Fez Primeira Eucaristia")
        c.setFillColor(colors.black)


# Casamento
    
    if ficha.casado_igreja:
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 600, f"Casado na Igreja - Data: {ficha.casado_igreja_data.strftime("%d/%m/%Y") if ficha.casado_igreja_data else ''}")
        c.setFont("Helvetica", 11)
        c.drawString(50, height - 615, f"Diocese: {ficha.casado_igreja_diocese}")
        c.drawString(50, height - 630, f"Paróquia: {ficha.casado_igreja_paroquia}")
        c.drawString(50, height - 645, f"Celebrante: {ficha.casado_igreja_celebrante} ")
    else:
        c.setFillColor(colors.red)
        c.setFont("Helvetica-Bold", 11)
        c.drawString(50, height - 600, f"Não casou na Igreja ")
        c.setFillColor(colors.black)

    c.setFont("Helvetica", 11)
    c.drawString(50, height - 700, f"Horário:")
    c.drawString(150, height - 700, f"{ficha.get_horario_display()}")
    
   
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome} <br/>CPF: {ficha.cpf}",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.showPage()  # Página 2
    

    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(100, height - 160, f"TERMO DE CONSENTIMENTO PARA TRATAMENTO ")
    c.drawString(120, height - 180, f"DE DADOS PESSOAIS SENSÍVEIS")
    


    paragrafo = Paragraph(
        f"Eu, {ficha.nome}, CPF {ficha.cpf}, por meio deste instrumento, <b>manifesto meu consentimento livre, informado e inequívoco para o tratamento dos meus dados pessoais"
        "</b>, nos seguintes termos, em conformidade com a Lei no 13.709/2018:"
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"1) Autorizo o tratamento a ser realizado pela PARÓQUIA NOSSA SENHORA APARECIDA, pessoa jurídica "
        f"de direito privado, inscrita no CNPJ no 44.802.999/0011-30, situada na Rua dois, no 349, Bairro Aparecida,"
        f"Rio Claro/SP, CEP 13500-270, e e-mail: pnsarc@hotmail.com, doravante denominada CONTROLADORA."
        ,style
    )
    frame = Frame(50, height - 460, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"2) Meus dados serão utilizados exclusivamente para:<br/>"
        f"    I) Inscrição e organização da Catequese;<br/>"
        f"    II) Emitir certificados de conclusão dos sacramentos;<br/>"
        f"    III) Permitir contato com pais e/ou responsáveis.<br/>"
        ,style
    )
    frame = Frame(50, height - 520, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"3) Autorizo o tratamento apenas dos seguintes dados pessoais: nome completo; data de nascimento;"
        f"naturalidade; documento de identificação (RG, CPF, certidão de nascimento ou equivalente); endereço completo; telefone de contato."
        f"Não haverá utilização dos dados para fins comerciais, divulgação indevida ou compartilhamento com terceiros estranhos às atividades religiosas."
        ,style
    )
    frame = Frame(50, height - 590, 500, 200)
    frame.addFromList([paragrafo], c)
    

    paragrafo = Paragraph(
        f"4) Declaro estar ciente de que o armazenamento ocorrerá:"
        f"I) Em fichas físicas e/ou sistemas informatizados da Paróquia;<br/>"
        f"II) Que o acesso será restrito a pessoas autorizadas (secretaria, coordenação da catequese, catequistas e o Pároco, quando necessário);<br/>"
        f"III) Que a CONTROLADORA adotará medidas técnicas e administrativas adequadas para proteger os dados contra acessos não autorizados, perda, divulgação indevida ou qualquer forma de tratamento inadequado.<br/>"
        ,style
    )
    frame = Frame(50, height - 670, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"5) Os dados serão mantidos enquanto estiver matriculado(a) na Catequese, pelo prazo máximo de 3 (três) anos, quando concluída as etapas sacramentais, respeitando-se a necessidade de registros paroquiais.<br/>"
        f"6) Estou ciente de que posso, a qualquer momento solicitar acesso aos meus dados pessoais; solicitar correção de dados incompletos, inexatos ou desatualizados; requerer a eliminação de dados desnecessários, excessivos ou tratados em desconformidade com a lei.<br/>"
        f"7) Declaro, ainda, estar ciente de que posso revogar este consentimento a qualquer momento, mediante solicitação formal e por escrito à Secretaria Paroquial.<br/>"
        ,style
    )
    frame = Frame(50, height - 780, 500, 200)
    frame.addFromList([paragrafo], c)
   
    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome} <br/>CPF: {ficha.cpf}",
        style
    )
    frame = Frame(250, height - 940, 500, 200)
    frame.addFromList([paragrafo], c)
    
    c.showPage()  # Página 3 
    
    c.drawImage(
        img_path,
        x=50,           # posição X
        y=height - 140, # posição Y (200px de altura da imagem)
        width=500,      # ajuste como preferir
        height=150,
        preserveAspectRatio=True,
        mask='auto'
    )
    c.setFont("Helvetica-Bold", 13)
    c.drawString(160, height - 160, f"AUTORIZAÇÃO PARA USO DE IMAGEM")
    
    styles = getSampleStyleSheet()
    style = styles["Normal"]
    style.fontName = "Helvetica"
    style.fontSize = 11
    style.leading = 15

    paragrafo = Paragraph(
        f"Eu, {ficha.nome}, CPF {ficha.cpf} "
        f", autorizo, de forma livre, expressa e informada, a Paróquia Nossa Senhora Aparecida, inscrita "
        f"no CNPJ sob o nº 44.802.999/0011-30, a utilizar a minha imagem, nome e voz, captados em fotografias e/ou vídeos durante atividades e eventos da paróquia, para fins de divulgação em meios impressos, digitais e redes sociais da paróquia, sem qualquer ônus."
        ,style
    )
    frame = Frame(50, height - 400, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Declaro estar ciente de que a utilização da minha imagem será feita de acordo com a Lei Geral de Proteção de Dados (Lei nº 13.709/2018 - LGPD), e que posso, a qualquer momento, revogar esta autorização mediante solicitação por escrito à paróquia. "
        ,style
    )
    frame = Frame(50, height - 480, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
        f"Estou ciente de que não tenho direito a qualquer remuneração pelo uso da minha imagem, nome e voz nos termos acima mencionados, e que a presente autorização é concedida por prazo indeterminado, podendo ser revogada a qualquer momento, mediante comunicação por escrito. "
        ,style
    )
    frame = Frame(50, height - 530, 500, 200)
    frame.addFromList([paragrafo], c)
    
    paragrafo = Paragraph(
        f"Por fim, declaro que a presente autorização foi feita de forma livre, sem qualquer coação, e que fui devidamente informado(a) sobre o tratamento dos dados pessoais do(a) menor pela paróquia. "
        ,style
    )
    frame = Frame(50, height - 580, 500, 200)
    frame.addFromList([paragrafo], c)

    paragrafo = Paragraph(
            f"Rio Claro, {data_hoje()}<br/><br/> _____________________________________________<br/> {ficha.nome} - CPF: {ficha.cpf}",
        style
    )
    frame = Frame(250, height - 660, 500, 200)
    frame.addFromList([paragrafo], c)
    
    
    c.save()

    return filename


from django.utils.timezone import localtime
import openpyxl
from .models import CatequeseInfantilModel, CrismaModel, Perseveranca_MEJ_Model, CatequeseAdultoModel
def gerar_Workbook():
    qs = CatequeseInfantilModel.objects.all().order_by('id')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Catequese Infantil"
    list_header = ['Nome','Nome do Pai','Nome da Mãe','Endereço','Cidade','UF','Celular do Pai','Celular da Mãe',]
    list_header += ['Batizado','Data do Batismo','Diocese do Batismo','Paróquia do Batismo','Celebrante do Batismo']
    list_header += ['Horário da Catequese']
    list_header += ['Possui Deficiência', 'Descrição da Deficiência', 'Possui Transtorno', 'Descrição do Transtorno']
    list_header += ['Medicamento de Uso Contínuo',  'Descrição do Medicamento', 'Horário do Medicamento']
    list_header += ['Acompanhamento Psicológico','Descrição do Acompanhamento',]
    list_header += ['Nome do Responsável','CPF do Responsável','Endereço do Responsável']
    list_header += ['Ficha Impressa','Ficha Assinada','Data de Registro']
    ws.append(list_header)
    for registro in qs:
        ws.append([
            registro.nome,
            registro.nome_pai,
            registro.nome_mae,
            registro.endereco,
            registro.cidade,
            registro.uf,
            registro.celular_pai,
            registro.celular_mae,
            'Sim' if registro.batizado else 'Não',
            registro.batizado_data.strftime('%d/%m/%Y') if registro.batizado_data else 'NÃO BATIZADO',
            registro.batizado_diocese,
            registro.batizado_paroquia,
            registro.batizado_celebrante,
            registro.get_horario_display(),
            'Sim' if registro.possui_deficiencia else 'Não',
            registro.descricao_deficiencia,
            'Sim' if registro.possui_transtorno else 'Não',
            registro.descricao_transtorno,
            'Sim' if registro.medicamento_uso_continuo else 'Não',
            registro.descricao_medicamento,
            registro.medicamento_horario,
            'Sim' if registro.acompanhamento_psicologico else 'Não',
            registro.descricao_acompanhamento,
            registro.nome_responsavel,
            registro.cpf_responsavel,
            registro.endereco_responsavel,
            "Sim" if registro.ficha_impressa else "Não",
            "Sim" if registro.ficha_assinada else "Não",    
            localtime(registro.criado_em).strftime('%d/%m/%Y - %H:%M:%S') if registro.criado_em else '',
        ])
    ws = wb.create_sheet(title="Crisma")
    
    qs = CrismaModel.objects.all().order_by('id')
    list_header = [
        'Nome',
        'Nome do Pai',
        'Nome da Mãe',
        'Endereço',
        'Cidade',
        'UF',
        'Celular do Pai',
        'Celular da Mãe',
]

    list_header += [
        'Batizado',
        'Data do Batismo',
        'Diocese do Batismo',
        'Paróquia do Batismo',
        'Celebrante do Batismo',
    ]

    list_header += [
        'Primeira Eucaristia',
        'Data da Primeira Eucaristia',
        'Diocese da Primeira Eucaristia',
        'Paróquia da Primeira Eucaristia',
        'Celebrante da Primeira Eucaristia',
    ]

    list_header += [
        'Horário da Crisma',
        'Nome do Padrinho',
        'Celular do Padrinho',
    ]

    list_header += [
        'Possui Deficiência',
        'Descrição da Deficiência',
        'Possui Transtorno',
        'Descrição do Transtorno',
    ]

    list_header += [
        'Medicamento de Uso Contínuo',
        'Descrição do Medicamento',
        'Horário do Medicamento',
    ]

    list_header += [
        'Acompanhamento Psicológico',
        'Descrição do Acompanhamento',
    ]

    list_header += [
        'Nome do Responsável',
        'CPF do Responsável',
        'Endereço do Responsável',
    ]

    list_header += [
        'Ficha Impressa',
        'Ficha Assinada',
        'Data de Registro',
    ]

    ws.append(list_header)

    for registro in qs:
        ws.append([
            registro.nome,
            registro.nome_pai,
            registro.nome_mae,
            registro.endereco,
            registro.cidade,
            registro.uf,
            registro.celular_pai,
            registro.celular_mae,

            'Sim' if registro.batizado else 'Não',
            registro.batizado_data.strftime('%d/%m/%Y') if registro.batizado_data else 'NÃO BATIZADO',
            registro.batizado_diocese,
            registro.batizado_paroquia,
            registro.batizado_celebrante,

            'Sim' if registro.primeira_eucaristia else 'Não',
            registro.primeira_eucaristia_data.strftime('%d/%m/%Y') if registro.primeira_eucaristia_data else 'NÃO REALIZADA',
            registro.primeira_eucaristia_diocese,
            registro.primeira_eucaristia_paroquia,
            registro.primeira_eucaristia_celebrante,

            registro.get_horario_display(),
            registro.padrinho_nome,
            registro.padrinho_celular,

            'Sim' if registro.possui_deficiencia else 'Não',
            registro.descricao_deficiencia,

            'Sim' if registro.possui_transtorno else 'Não',
            registro.descricao_transtorno,

            'Sim' if registro.medicamento_uso_continuo else 'Não',
            registro.descricao_medicamento,
            registro.medicamento_horario,

            'Sim' if registro.acompanhamento_psicologico else 'Não',
            registro.descricao_acompanhamento,

            registro.nome_responsavel,
            registro.cpf_responsavel,
            registro.endereco_responsavel,

            'Sim' if registro.ficha_impressa else 'Não',
            'Sim' if registro.ficha_assinada else 'Não',

            localtime(registro.criado_em).strftime('%d/%m/%Y - %H:%M:%S') if registro.criado_em else '',
        ])
    
    ws = wb.create_sheet(title="Perseverança MEJ")
    
    qs = Perseveranca_MEJ_Model.objects.all().order_by('id')
    list_header = [
        'Nome',
        'Nome do Pai',
        'Nome da Mãe',
        'Endereço',
        'Cidade',
        'UF',
        'Celular do Pai',
        'Celular da Mãe',
    ]

    list_header += [
        'Batizado',
        'Data do Batismo',
        'Diocese do Batismo',
        'Paróquia do Batismo',
        'Celebrante do Batismo',
    ]

    list_header += [
        'Primeira Eucaristia',
        'Data da Primeira Eucaristia',
        'Diocese da Primeira Eucaristia',
        'Paróquia da Primeira Eucaristia',
        'Celebrante da Primeira Eucaristia',
    ]

    list_header += [
        'Horário da Perseverança / MEJ',
    ]

    list_header += [
        'Possui Deficiência',
        'Descrição da Deficiência',
        'Possui Transtorno',
        'Descrição do Transtorno',
    ]

    list_header += [
        'Medicamento de Uso Contínuo',
        'Descrição do Medicamento',
        'Horário do Medicamento',
    ]

    list_header += [
        'Acompanhamento Psicológico',
        'Descrição do Acompanhamento',
    ]

    list_header += [
        'Nome do Responsável',
        'CPF do Responsável',
        'Endereço do Responsável',
    ]

    list_header += [
        'Ficha Impressa',
        'Ficha Assinada',
        'Data de Registro',
    ]

    ws.append(list_header)

    for registro in qs:
        ws.append([
            registro.nome,
            registro.nome_pai,
            registro.nome_mae,
            registro.endereco,
            registro.cidade,
            registro.uf,
            registro.celular_pai,
            registro.celular_mae,

            'Sim' if registro.batizado else 'Não',
            registro.batizado_data.strftime('%d/%m/%Y') if registro.batizado_data else 'NÃO BATIZADO',
            registro.batizado_diocese,
            registro.batizado_paroquia,
            registro.batizado_celebrante,

            'Sim' if registro.primeira_eucaristia else 'Não',
            registro.primeira_eucaristia_data.strftime('%d/%m/%Y') if registro.primeira_eucaristia_data else 'NÃO REALIZADA',
            registro.primeira_eucaristia_diocese,
            registro.primeira_eucaristia_paroquia,
            registro.primeira_eucaristia_celebrante,

            registro.get_horario_display(),

            'Sim' if registro.possui_deficiencia else 'Não',
            registro.descricao_deficiencia,

            'Sim' if registro.possui_transtorno else 'Não',
            registro.descricao_transtorno,

            'Sim' if registro.medicamento_uso_continuo else 'Não',
            registro.descricao_medicamento,
            registro.medicamento_horario,

            'Sim' if registro.acompanhamento_psicologico else 'Não',
            registro.descricao_acompanhamento,

            registro.nome_responsavel,
            registro.cpf_responsavel,
            registro.endereco_responsavel,

            'Sim' if registro.ficha_impressa else 'Não',
            'Sim' if registro.ficha_assinada else 'Não',

            localtime(registro.criado_em).strftime('%d/%m/%Y - %H:%M:%S') if registro.criado_em else '',
        ])
    ws = wb.create_sheet(title="Catequese Adulto")
    qs = CatequeseAdultoModel.objects.all().order_by('id')
    
    list_header = [
        'Nome',
        'CPF',
        'Celular',
        'Nome do Pai',
        'Nome da Mãe',
        'Endereço',
        'Cidade',
        'UF',
        'Estado Civil',
    ]

    list_header += [
        'Batizado',
        'Data do Batismo',
        'Diocese do Batismo',
        'Paróquia do Batismo',
        'Celebrante do Batismo',
    ]

    list_header += [
        'Primeira Eucaristia',
        'Data da Primeira Eucaristia',
        'Diocese da Primeira Eucaristia',
        'Paróquia da Primeira Eucaristia',
        'Celebrante da Primeira Eucaristia',
    ]

    list_header += [
        'Casado na Igreja',
        'Data do Casamento',
        'Diocese do Casamento',
        'Paróquia do Casamento',
        'Celebrante do Casamento',
    ]

    list_header += [
        'Horário da Catequese (Adulto)',
        'Nome do Padrinho',
        'Celular do Padrinho',
    ]

    list_header += [
        'Ficha Impressa',
        'Ficha Assinada',
        'Data de Registro',
    ]

    ws.append(list_header)

    for registro in qs:
        ws.append([
            registro.nome,
            registro.cpf,
            registro.celular,
            registro.nome_pai,
            registro.nome_mae,
            registro.endereco,
            registro.cidade,
            registro.uf,
            registro.estado_civil,

            'Sim' if registro.batizado else 'Não',
            registro.batizado_data.strftime('%d/%m/%Y') if registro.batizado_data else 'NÃO BATIZADO',
            registro.batizado_diocese,
            registro.batizado_paroquia,
            registro.batizado_celebrante,

            'Sim' if registro.primeira_eucaristia else 'Não',
            registro.primeira_eucaristia_data.strftime('%d/%m/%Y') if registro.primeira_eucaristia_data else 'NÃO REALIZADA',
            registro.primeira_eucaristia_diocese,
            registro.primeira_eucaristia_paroquia,
            registro.primeira_eucaristia_celebrante,

            'Sim' if registro.casado_igreja else 'Não',
            registro.casado_igreja_data.strftime('%d/%m/%Y') if registro.casado_igreja_data else 'NÃO INFORMADO',
            registro.casado_igreja_diocese,
            registro.casado_igreja_paroquia,
            registro.casado_igreja_celebrante,

            registro.get_horario_display(),
            registro.padrinho_nome,
            registro.padrinho_celular,

            'Sim' if registro.ficha_impressa else 'Não',
            'Sim' if registro.ficha_assinada else 'Não',

            localtime(registro.criado_em).strftime('%d/%m/%Y - %H:%M:%S') if registro.criado_em else '',
        ])
    return wb